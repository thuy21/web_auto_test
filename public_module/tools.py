from xlutils.copy import copy
import xlrd
from openpyxl import load_workbook


def xlsx_read(dir, sheet):
    """读取exel文件"""
    list = []
    wb = load_workbook(dir)
    ws = wb[sheet]
    for i in range(2, ws.max_row + 1):
        list_item = []
        for j in range(1, ws.max_column + 1):
            if ws.cell(i, j).value is not None:
                list_item.append(ws.cell(i, j).value)
        list.append(list_item)
    return list


def xls_update(dir, list, row):
    """修改excl文件"""
    book1 = xlrd.open_workbook(dir, formatting_info=True)
    book2 = copy(book1)
    sheet = book2.get_sheet(0)
    for item in list:
        for i in range(0, len(item)):
            if item[i] == '无':
                i += 1
                continue
            sheet.write(row, i, item[i])
        row += 1

    book2.save(dir)