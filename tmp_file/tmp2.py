import openpyxl
from openpyxl import load_workbook

exel = load_workbook('tmp.xlsx')

page1 = exel['Sheet1']

with open('res.txt') as f:
    while True:
        res = f.readline()
        if len(res) == 0:
            break
        num = int(res)
        print(res)
        page1.cell(num,1).value = "***"

exel.save('tmp2.xlsx')